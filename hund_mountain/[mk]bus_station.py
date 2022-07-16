import googlemaps
import openpyxl
import os
from dotenv import load_dotenv
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from time import sleep


filename = 'source/bus_station.xlsx'
wb3 = openpyxl.Workbook()
sheet = wb3.active


target_url = 'https://www.kobus.co.kr/ugd/trmlgd/Trmlgd.do'

service = Service('../../drivers/chromedriver.exe')
driver = webdriver.Chrome(service=service)

sleep(3)
driver.get(target_url)
sleep(3)

soup = BeautifulSoup(driver.page_source, 'html.parser')
# soup는 객체 (html.parser가 문서를 객체로 바꿔주었다)


area_lst = ['서울', '경기', '인천', '강원', '충남', '대전', '충북', '전북', '경북', '대구', '광주', '전남', '경남', '울산', '부산']
r = 1

for i in range(2, 17):
    select = driver.find_element_by_xpath('/html/body/div[1]/div[4]/div[2]/div[3]/div[2]/div/div/div/div[2]')
    select.click()
    sleep(2)
    area = driver.find_element_by_xpath(f'/html/body/div[1]/div[4]/div[2]/div[3]/div[2]/div/div/div/div[3]/div/ul/li[{i}]')
    area.click()
    sleep(2)

    tms = soup.select(f'tr[data-label={area_lst[i-2]}]')

    for tm in tms:
        tm_nm = tm.select_one("th").text
        tm_addr = tm.select_one("td").text

        sheet.cell(r, 1).value = r
        sheet.cell(r, 2).value = tm_nm
        sheet.cell(r, 3).value = tm_addr
        r += 1

print(r)

#버스터미널 위경도 추가

load_dotenv()
google_key = os.environ.get("GOOGLE_API_KEY")
gmaps = googlemaps.Client(key=google_key)

j = 0
for i in range(r):
    j += 1
    try :
        geocode_result = gmaps.geocode(sheet.cell(j, 3).value)
        sheet.cell(j, 4). value = geocode_result[0]['geometry']['location']['lat']
        sheet.cell(j, 5). value = geocode_result[0]['geometry']['location']['lng']
    except:
        continue

sheet.insert_rows(1, 1)  # 제일 처음에 빈 1행 추가
# no, tm_nm, addr, lat, lot

wb3.save(filename)